import customtkinter
from openpyxl import load_workbook

def verificar_usuario_existente(usuario):
    workbook = load_workbook('C:/Users/Administrador/Desktop/Projetohelton/cadastro_usuarios.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        if row[0] == usuario:
            return True

    return False

def criar_login_senha(usuario, senha):
    if verificar_usuario_existente(usuario):
        resultado_label.configure(text="Usuário já cadastrado. Por favor, escolha outro nome de usuário.")
    else:
        workbook = load_workbook('C:/Users/Administrador/Desktop/Projetohelton/cadastro_usuarios.xlsx')
        sheet = workbook.active

        row = sheet.max_row + 1
        sheet.cell(row=row, column=1).value = usuario
        sheet.cell(row=row, column=2).value = senha

        workbook.save('C:/Users/Administrador/Desktop/Projetohelton/cadastro_usuarios.xlsx')
        resultado_label.configure(text="Cadastro realizado com sucesso!")

def cadastrar_usuario():
    usuario = usuario_entry.get()
    senha = senha_entry.get()

    criar_login_senha(usuario, senha)

def voltar_menu():
    usuario_entry.delete(0, 'end')
    senha_entry.delete(0, 'end')

    import testeMenuPrincipal

janela = customtkinter.CTk()
janela.geometry("400x300")

frame = customtkinter.CTkFrame(janela)
frame.pack(padx=10, pady=10)

usuario_label = customtkinter.CTkLabel(frame, text="Nome de Usuário:")
usuario_label.pack()

usuario_entry = customtkinter.CTkEntry(frame)
usuario_entry.pack(pady=5)

senha_label = customtkinter.CTkLabel(frame, text="Senha:")
senha_label.pack()

senha_entry = customtkinter.CTkEntry(frame, show="*")
senha_entry.pack(pady=5)

botao_cadastrar = customtkinter.CTkButton(janela, text="Cadastrar", command=cadastrar_usuario)
botao_cadastrar.pack(pady=10)

botao_voltar = customtkinter.CTkButton(janela, text="Voltar", command=voltar_menu)
botao_voltar.pack(pady=10)

resultado_label = customtkinter.CTkLabel(janela)
resultado_label.pack(pady=10)

janela.mainloop()
